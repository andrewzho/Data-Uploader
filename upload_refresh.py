"""
upload_refresh.py

Simple uploader + SQL runner for the workspace SQL scripts.
- Creates inbound/<script-folder> for each .sql in cwd (ordered by name)
- Writes a template config.json mapping folder -> target_table (edit before uploading)
- Uploads Excel files (*.xlsx, *.xls) from those folders into the configured table (fast_executemany via pyodbc)

Usage:
  python upload_refresh.py --init         # create folders + config.json
  python upload_refresh.py --upload       # upload files from configured folders
  python upload_refresh.py --run-sql      # execute SQL scripts in order
  python upload_refresh.py --upload --run-sql
  python upload_refresh.py --list-tables   # list accessible base tables via INFORMATION_SCHEMA
  python upload_refresh.py --test-connection  # test DB connection and report server/user info

Dependencies: pandas, pyodbc, openpyxl
Install: pip install pandas pyodbc openpyxl
"""

from pathlib import Path
import os
import json
import argparse
import re
import sys
import tempfile
import csv
from datetime import date as date_type

try:
    import pandas as pd
except Exception:
    pd = None

try:
    import pyodbc
except Exception:
    pyodbc = None

try:
    from difflib import SequenceMatcher
except Exception:
    SequenceMatcher = None

try:
    import numpy as np
except Exception:
    np = None


def convert_numpy_to_python(val):
    """
    Convert numpy types to Python native types for pyodbc compatibility.
    pyodbc doesn't handle numpy.int64, numpy.float64, etc. directly.
    """
    if np is not None:
        # Check if it's a numpy scalar type
        if isinstance(val, (np.integer, np.floating)):
            # Convert numpy integer/float to Python native type
            if isinstance(val, np.integer):
                return int(val)
            elif isinstance(val, np.floating):
                return float(val)
        # Check for numpy datetime64
        elif isinstance(val, np.datetime64):
            # Convert to pandas Timestamp (which pyodbc can handle) or Python datetime
            return pd.Timestamp(val)
    # Check by type name as fallback (in case numpy isn't imported but types are still numpy)
    type_name = type(val).__name__
    if type_name.startswith('int') and hasattr(val, 'item'):
        # numpy integer type
        return int(val.item())
    elif type_name.startswith('float') and hasattr(val, 'item'):
        # numpy float type
        return float(val.item())
    # Return as-is if not a numpy type
    return val


def list_sql_files(base: Path):
    return sorted([p.name for p in base.glob('*.sql')])


def safe_dirname(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', '_', name)


def create_inbound_dirs(sql_files, base: Path):
    # Ensure the base inbound directory exists (base is typically project/inbound)
    base.mkdir(parents=True, exist_ok=True)
    folders = []
    for f in sql_files:
        stem = Path(f).stem
        d = base / safe_dirname(stem)
        d.mkdir(parents=True, exist_ok=True)
        # Return relative folder path (relative to project root) so the generated
        # `config.json` is portable across machines. e.g. 'inbound/1 - ScriptName'
        try:
            rel = str(d.relative_to(Path(__file__).parent.resolve()))
        except Exception:
            # Fallback: use a path under 'inbound'
            rel = os.path.join('inbound', safe_dirname(stem))
        folders.append(rel)
    return folders


def write_template_config(sql_files, folders, path: Path):
    cfg = {
        "db": {
            "driver": "ODBC Driver 17 for SQL Server",
            "server": "SERVER_NAME_OR_HOST",
            "database": "NYPCSQL01",
            "trusted_connection": True,
            "username": "",
            "password": ""
        },
        "folders": []
    }
    for sf, fd in zip(sql_files, folders):
        # Ensure folder entries in the template are relative paths where possible
        folder_value = fd
        p = Path(fd)
        if p.is_absolute():
            try:
                folder_value = str(p.relative_to(Path(__file__).parent.resolve()))
            except Exception:
                # leave absolute path if we can't relativize it
                folder_value = fd

        cfg["folders"].append({
            "script": sf,
            "folder": folder_value,
            "target_table": None,
            "file_patterns": ["*.xlsx", "*.xls"],
            "upload_mode": "append"  # options: 'append', 'delete'
        })
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(cfg, fh, indent=2)
    print(f"Template config written to {path}. Edit target_table entries before uploading.")


def connect_from_cfg(dbcfg: dict):
    if pyodbc is None:
        raise SystemExit("pyodbc is not installed. Install with: pip install pyodbc")
    driver = dbcfg.get('driver', 'ODBC Driver 17 for SQL Server')
    server = dbcfg.get('server')
    database = dbcfg.get('database')
    trusted = dbcfg.get('trusted_connection', True)
    if not server or not database:
        raise ValueError('server and database must be set in config db section')
    if trusted:
        conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection=yes;"
        return pyodbc.connect(conn_str, autocommit=False)
    else:
        user = dbcfg.get('username')
        pwd = dbcfg.get('password')
        if not user or not pwd:
            raise ValueError('username and password required when trusted_connection is False')
        conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={user};PWD={pwd}"
        return pyodbc.connect(conn_str, autocommit=False)


def test_connection(cfg_path: Path):
    """Attempt a DB connection using config and print basic server/user info."""
    cfg = json.load(open(cfg_path, 'r', encoding='utf-8'))
    try:
        conn = connect_from_cfg(cfg['db'])
    except Exception as e:
        print('Connection failed:', e)
        return 1
    try:
        cur = conn.cursor()
        # Fetch current login, server name and version
        # use safe alias names (avoid reserved keywords like current_user)
        cur.execute("SELECT SUSER_SNAME() AS current_login, @@SERVERNAME AS server_name, @@VERSION AS version")
        row = cur.fetchone()
        if row:
            print('Connected successfully')
            # access by alias names returned by the query
            try:
                print('Current user:', row.current_login)
                print('Server name :', row.server_name)
                print('Version     :', row.version.split('\n')[0])
            except Exception:
                # fallback to tuple-style indexing if attribute names differ
                print('Current user:', row[0])
                print('Server name :', row[1])
                print('Version     :', str(row[2]).split('\n')[0])
        else:
            print('Connected but no info returned')
        # Optionally try a simple metadata query to ensure permissions
        try:
            cur.execute("SELECT COUNT(*) AS table_count FROM INFORMATION_SCHEMA.TABLES")
            rc = cur.fetchone()
            if rc:
                print('Information schema tables visible:', rc.table_count)
        except Exception as me:
            print('Warning: could not query INFORMATION_SCHEMA.TABLES:', me)
        return 0
    finally:
        conn.close()


def get_tables_list(cfg_path: Path):
    """Get list of accessible base tables as list of (schema, table_name, full_name) tuples"""
    cfg = json.load(open(cfg_path, 'r', encoding='utf-8'))
    try:
        conn = connect_from_cfg(cfg['db'])
    except Exception as e:
        return []
    try:
        cur = conn.cursor()
        cur.execute("SELECT TABLE_SCHEMA, TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' ORDER BY TABLE_SCHEMA, TABLE_NAME")
        rows = cur.fetchall()
        tables = []
        for r in rows:
            try:
                schema = r.TABLE_SCHEMA
                table = r.TABLE_NAME
            except Exception:
                # fallback tuple
                schema = r[0]
                table = r[1]
            # Create full name: schema.table or [database].schema.table if database is specified
            db = cfg.get('db', {}).get('database', '')
            if db:
                full_name = f"[{db}].[{schema}].[{table}]"
            else:
                full_name = f"[{schema}].[{table}]"
            tables.append((schema, table, full_name))
        return tables
    finally:
        conn.close()


def list_tables(cfg_path: Path):
    """List accessible base tables via INFORMATION_SCHEMA.TABLES"""
    tables = get_tables_list(cfg_path)
    if not tables:
        print('No base tables found or insufficient permissions.')
        return 1
    print('Accessible tables:')
    for schema, table, full_name in tables:
        print(full_name)
    return 0


def split_sql_batches(sql_text: str):
    # Split on lines that contain only GO (case-insensitive)
    batches = re.split(r'(?im)^\s*GO\s*;?\s*$', sql_text)
    return [b.strip() for b in batches if b.strip()]
def upload_excel_in_chunks(file_path, conn, table, table_cols, upload_mode='append', chunk_size=25000, log_callback=None):
    """
    Read and upload Excel or CSV file in chunks to avoid loading entire file into memory.
    This is much more memory-efficient for large files.
    
    Args:
        file_path: Path to Excel or CSV file
        conn: Database connection
        table: Target table name
        table_cols: Table column metadata
        upload_mode: 'append' or 'delete'
        chunk_size: Number of rows to process at a time
        log_callback: Optional function to call with log messages
    
    Returns:
        Total number of rows uploaded
    """
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg, flush=True)
    
    import time
    file_ext = Path(file_path).suffix.lower()
    
    # First, get total row count and column names
    log(f"  Analyzing file structure...")
    
    if file_ext == '.csv':
        # For CSV, read first row to get headers and count total rows
        df_sample = pd.read_csv(file_path, nrows=1)
        headers = list(df_sample.columns)
        
        # Count total rows (this is fast for CSV)
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            total_rows = sum(1 for line in f) - 1  # Subtract header row
    else:
        # Excel file - use openpyxl for efficient reading
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Get column headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value is not None else f"Column{len(headers)+1}")
        
        total_rows = ws.max_row - 1  # Subtract header row
        wb.close()  # Close read-only workbook
    
    log(f"  File has {total_rows:,} rows, {len(headers)} columns")
    
    # Handle upload mode
    log(f"  Upload mode: {upload_mode}")
    cursor = conn.cursor()
    if upload_mode == 'delete':
        try:
            log(f"  Deleting existing data from table (upload_mode='delete')...")
            cursor.execute(f"DELETE FROM {table}")
            conn.commit()  # Commit the delete before starting inserts
            log(f"  ✓ Cleared existing data from table")
        except Exception as e:
            log(f"  ✗ Warning: Could not delete table data: {e}")
            conn.rollback()
    else:
        log(f"  Appending to existing data (upload_mode='append')")
    
    # Read and process in chunks
    total_uploaded = 0
    chunk_num = 0
    start_row = 2  # Start at row 2 (skip header row 1)
    
    log(f"  Starting chunked processing (reading {chunk_size:,} rows at a time)...")
    
    while start_row <= total_rows + 1:  # +1 because max_row includes header
        chunk_num += 1
        try:
            # Calculate how many rows to read in this chunk
            rows_to_read = min(chunk_size, total_rows + 2 - start_row)
            
            if rows_to_read <= 0:
                break  # No more data
            
            # Read chunk using skiprows and nrows
            chunk_start = time.time()
            log(f"  Reading chunk {chunk_num} (rows {start_row}-{start_row + rows_to_read - 1})...")
            
            # Read chunk based on file type
            skip_list = list(range(0, start_row - 1))  # Skip header (0) and rows before start_row
            
            if file_ext == '.csv':
                # For CSV, use skiprows and nrows
                df_chunk = pd.read_csv(
                    file_path,
                    skiprows=skip_list,
                    nrows=rows_to_read
                )
            else:
                # Excel file
                df_chunk = pd.read_excel(
                    file_path, 
                    engine='openpyxl',
                    skiprows=skip_list,  # List of row indices to skip (0-indexed)
                    nrows=rows_to_read
                )
            
            if df_chunk.empty:
                break  # No more data
            
            # Set column names from headers we extracted
            if len(df_chunk.columns) == len(headers):
                df_chunk.columns = headers
            elif len(df_chunk.columns) > 0:
                # If column count doesn't match, use what we have
                log(f"  Warning: Column count mismatch (expected {len(headers)}, got {len(df_chunk.columns)})")
            
            # Prepare this chunk
            from upload_refresh import prepare_dataframe_for_table
            df_prepared = prepare_dataframe_for_table(df_chunk, table_cols, filename=Path(file_path).name)
            
            # Upload this chunk using regular method
            # For chunks after the first, always append (table was already cleared at start if needed)
            chunk_upload_mode = 'append' if chunk_num > 1 else upload_mode
            upload_df_to_table(conn, df_prepared, table, upload_mode=chunk_upload_mode, table_cols=table_cols)
            
            # CRITICAL: Commit after each chunk to avoid huge transaction log and performance degradation
            conn.commit()
            
            total_uploaded += len(df_prepared)
            chunk_time = time.time() - chunk_start
            progress_pct = (total_uploaded / total_rows) * 100 if total_rows > 0 else 0
            
            log(f"  ✓ Chunk {chunk_num}: {len(df_prepared):,} rows uploaded ({total_uploaded:,}/{total_rows:,}, {progress_pct:.1f}%) in {chunk_time:.1f}s")
            
            # Move to next chunk
            start_row += rows_to_read
            
            # Free memory explicitly
            del df_chunk
            del df_prepared
            import gc
            gc.collect()  # Force garbage collection to free memory
            
        except Exception as e:
            log(f"  ✗ Error processing chunk {chunk_num}: {e}")
            raise
    
    # Final commit (though we already commit after each chunk)
    try:
        conn.commit()
    except:
        pass  # May already be committed
    log(f"  ✓ Chunked processing complete: {total_uploaded:,} total rows uploaded")
    return total_uploaded


def upload_df_to_table(conn, df, table, upload_mode='append', table_cols=None):
    """
    Upload DataFrame to SQL Server table.
    
    upload_mode options:
    - 'append': Add data without clearing existing data
    - 'delete': Delete all existing data, then insert new data (uses DELETE instead of TRUNCATE)
    """
    cursor = conn.cursor()
    cols = list(df.columns)
    if not cols:
        print('No columns found in file, skipping upload to', table)
        return
    
    # Handle table clearing based on upload_mode
    if upload_mode == 'delete':
        try:
            # DELETE is slower but works with foreign keys (can be rolled back)
            cursor.execute(f"DELETE FROM {table}")
        except Exception as e:
            print(f"Warning: Could not delete table data: {e}")
    elif upload_mode == 'truncate':
        # Legacy support for 'truncate' mode
        try:
            cursor.execute(f"TRUNCATE TABLE {table}")
        except Exception as e:
            # Fallback to DELETE if TRUNCATE fails (e.g., due to foreign keys)
            print(f"TRUNCATE failed, attempting DELETE: {e}")
            cursor.execute(f"DELETE FROM {table}")
    # 'append' mode does not clear the table
    
    placeholders = ", ".join("?" for _ in cols)
    col_list = ", ".join(f"[{c}]" for c in cols)
    sql = f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"
    
    # Build a map of column max lengths from table metadata if available
    col_max_lengths = {}
    if table_cols:
        for col_name, sql_type, max_length in table_cols:
            # Extract max length from SQL type (e.g., "NVARCHAR(255)" -> 255)
            if max_length and max_length > 0:
                col_max_lengths[col_name] = max_length
            elif 'VARCHAR' in sql_type.upper() or 'NVARCHAR' in sql_type.upper():
                # Try to extract from type string if max_length not provided
                match = re.search(r'\((\d+)\)', sql_type)
                if match:
                    col_max_lengths[col_name] = int(match.group(1))
    
    # Build a map of column data types for proper conversion
    col_data_types = {}
    col_is_date_only = {}  # Track which columns are DATE (not DATETIME)
    col_is_datetime = {}  # Track which columns are DATETIME/DATETIME2 (keep time component)
    # Cache date range constants to avoid recreating on every row
    min_date = date_type(1, 1, 1)  # SQL Server DATE minimum
    max_date = date_type(9999, 12, 31)  # SQL Server DATE maximum
    min_datetime = pd.Timestamp('1753-01-01')  # SQL Server DATETIME minimum
    max_datetime = pd.Timestamp('9999-12-31 23:59:59')  # SQL Server DATETIME maximum
    if table_cols:
        for col_name, sql_type, _ in table_cols:
            sql_type_lower = sql_type.lower()
            col_data_types[col_name] = sql_type_lower
            # Check if it's DATE (not DATETIME/DATETIME2) - DATE columns need date-only values
            col_is_date_only[col_name] = (sql_type_lower == 'date')
            # Check if it's DATETIME/DATETIME2/SMALLDATETIME - these keep time component
            col_is_datetime[col_name] = (sql_type_lower in ('datetime', 'datetime2', 'smalldatetime'))
    
    # OPTIMIZATION: Pre-process DATE and DATETIME columns using vectorized operations
    # This is MUCH faster than row-by-row conversion (10-100x speedup)
    df_processed = df.copy()
    
    # Process DATE columns in bulk (vectorized) - convert to date objects
    for col_name in df_processed.columns:
        if col_name in col_is_date_only and col_is_date_only[col_name]:
            try:
                # Convert to datetime first (handles strings, timestamps, etc.)
                dt_series = pd.to_datetime(df_processed[col_name], errors='coerce')
                # Extract date part (returns date objects)
                date_series = dt_series.dt.date
                # Replace out-of-range dates and NaT with None
                mask = pd.isna(dt_series) | (date_series < min_date) | (date_series > max_date)
                date_series.loc[mask] = None
                df_processed[col_name] = date_series
            except Exception:
                # Fallback: leave as-is, will be handled row-by-row
                pass
    
    # Process DATETIME columns in bulk (vectorized) - keep as Timestamp, convert to Python datetime row-by-row
    for col_name in df_processed.columns:
        if col_name in col_is_datetime and col_is_datetime[col_name]:
            try:
                # Convert to datetime (handles strings, etc.)
                dt_series = pd.to_datetime(df_processed[col_name], errors='coerce')
                # Replace out-of-range datetimes with NaT
                mask = (dt_series < min_datetime) | (dt_series > max_datetime)
                dt_series.loc[mask] = pd.NaT
                df_processed[col_name] = dt_series
            except Exception:
                # Fallback: leave as-is, will be handled row-by-row
                pass
    
    # Convert DataFrame to list of tuples, converting pandas NA/NaN to None for SQL Server
    # Use itertuples() instead of iterrows() to better preserve data types, especially booleans
    # This is more memory efficient and faster than iterrows()
    total_rows = len(df_processed)
    if total_rows > 50000:
        print(f"Converting {total_rows:,} rows to upload format...", end=' ', flush=True)
    data = []
    rows_processed = 0
    for row_tuple in df_processed.itertuples(index=False):
        row_data = []
        for i, val in enumerate(row_tuple):
            col_name = cols[i]
            # Special handling for BIT columns - ensure True/False are preserved
            if col_name in col_data_types and 'bit' in col_data_types[col_name]:
                # Handle pandas NA/NaN first
                if pd.isna(val):
                    row_data.append(None)
                else:
                    # Convert numpy types to Python native types first
                    val = convert_numpy_to_python(val)
                    
                    # Force conversion to Python bool to preserve True/False
                    # iterrows/itertuples might convert to int, so explicitly convert back
                    if isinstance(val, bool):
                        # Already a bool - keep as is
                        row_data.append(val)
                    elif isinstance(val, (int, float)) and val in (0, 1):
                        # Convert 1/0 back to True/False
                        row_data.append(True if val == 1 else False)
                    else:
                        # Try to convert to bool
                        try:
                            # Use explicit bool() conversion to ensure True/False
                            bool_val = bool(val)
                            row_data.append(bool_val)
                        except:
                            row_data.append(None)
            else:
                # Handle pandas NA (from nullable dtypes like Int64, boolean, string)
                if pd.isna(val):
                    row_data.append(None)
                else:
                    # Convert numpy types to Python native types for pyodbc compatibility
                    val = convert_numpy_to_python(val)
                    
                    # DATE and DATETIME columns are already pre-processed above using vectorized operations
                    # Just need to handle edge cases and convert Timestamp to Python datetime for DATETIME
                    if col_name in col_is_datetime and col_is_datetime[col_name]:
                        # DATETIME columns: convert Timestamp to Python datetime
                        if isinstance(val, pd.Timestamp):
                            if pd.isna(val):
                                val = None
                            else:
                                try:
                                    val = val.to_pydatetime()
                                except (ValueError, OverflowError, AttributeError):
                                    val = None
                    
                    # Truncate string values if they exceed column max length
                    if isinstance(val, str) and col_name in col_max_lengths:
                        max_len = col_max_lengths[col_name]
                        if len(val) > max_len:
                            val = val[:max_len]
                            print(f"Warning: Truncated '{col_name}' value (length {len(val)} > {max_len})")
                    row_data.append(val)
        data.append(tuple(row_data))
        rows_processed += 1
        # Show progress for very large files
        if total_rows > 50000 and rows_processed % 10000 == 0:
            progress_pct = (rows_processed / total_rows) * 100
            print(f"\rConverting {total_rows:,} rows to upload format... {rows_processed:,}/{total_rows:,} ({progress_pct:.1f}%)", end='', flush=True)
    
    if total_rows > 50000:
        print()  # New line after progress updates
    else:
        print("✓", flush=True)
    
    # Final check: Ensure BIT columns are True/False, not 1/0
    # This is a safety check in case any values slipped through as integers
    if table_cols:
        for col_name, sql_type, _ in table_cols:
            if 'bit' in sql_type.lower() and col_name in cols:
                col_idx = cols.index(col_name)
                # Rebuild data list with corrected boolean values
                corrected_data = []
                for row_data in data:
                    if col_idx < len(row_data):
                        val = row_data[col_idx]
                        if isinstance(val, (int, float)) and val in (0, 1):
                            # Convert 1/0 to True/False
                            row_data_list = list(row_data)
                            row_data_list[col_idx] = True if val == 1 else False
                            corrected_data.append(tuple(row_data_list))
                        else:
                            corrected_data.append(row_data)
                    else:
                        corrected_data.append(row_data)
                data = corrected_data
                break  # Only need to check once per column
    
    cursor.fast_executemany = True
    
    # Process in batches to avoid memory errors with large datasets
    # Reduced batch size for better performance and more frequent commits
    batch_size = 5000  # Reduced from 10,000 for better performance
    total_rows = len(data)
    
    try:
        if total_rows > batch_size:
            print(f"Large dataset detected ({total_rows:,} rows). Processing in batches of {batch_size:,}...", flush=True)
            rows_uploaded = 0
            total_batches = (total_rows + batch_size - 1) // batch_size
            for i in range(0, total_rows, batch_size):
                batch = data[i:i + batch_size]
                batch_num = (i // batch_size) + 1
                print(f"  Uploading batch {batch_num}/{total_batches} ({len(batch):,} rows)...", end=' ', flush=True)
                cursor.executemany(sql, batch)
                # CRITICAL: Commit after each batch to avoid transaction log growth
                conn.commit()
                rows_uploaded += len(batch)
                progress_pct = (rows_uploaded / total_rows) * 100
                print(f"✓ ({rows_uploaded:,}/{total_rows:,} rows, {progress_pct:.1f}%)", flush=True)
            print(f"✓ All {total_rows:,} rows uploaded successfully!", flush=True)
        else:
            # Small dataset - upload all at once
            print(f"Uploading {total_rows:,} rows...", end=' ', flush=True)
            cursor.executemany(sql, data)
            conn.commit()
            print(f"✓ Complete!", flush=True)
    except Exception as e:
        # Extract detailed error information from pyodbc errors
        error_details = {}
        error_msg = str(e)
        
        # Get the full exception representation
        import traceback
        full_traceback = traceback.format_exc()
        
        # Try multiple ways to extract error information
        # Method 1: Direct string representation
        error_str_full = str(e)
        
        # Method 2: Try pyodbc error format
        if hasattr(e, 'args') and len(e.args) > 0:
            # pyodbc errors often have args[0] as a tuple or string
            first_arg = e.args[0]
            if isinstance(first_arg, tuple) and len(first_arg) >= 2:
                # pyodbc error format: (error_code, error_message)
                error_details['error_code'] = first_arg[0] if len(first_arg) > 0 else None
                error_details['error_message'] = first_arg[1] if len(first_arg) > 1 else str(e)
                error_str_full = str(first_arg[1]) if len(first_arg) > 1 else error_str_full
            elif isinstance(first_arg, str):
                error_details['error_message'] = first_arg
                error_str_full = first_arg
            else:
                error_details['error_message'] = str(first_arg)
                error_str_full = str(first_arg)
        
        # Method 3: Check for pyodbc-specific attributes
        if hasattr(e, 'value'):
            error_str_full = str(e.value) if error_str_full == str(e) else error_str_full
        
        # Extract SQL Server error codes and line numbers from error string
        sql_error_match = re.search(r'\[(\d+)\]', error_str_full)
        if sql_error_match:
            error_details['sql_error_code'] = sql_error_match.group(1)
        
        line_match = re.search(r'\((\d+)\)', error_str_full)
        if line_match:
            error_details['line_number'] = line_match.group(1)
        
        # Look for common SQL error patterns
        truncation_match = re.search(r'String or binary data would be truncated.*?column \'([^\']+)\'', error_str_full, re.IGNORECASE)
        if truncation_match:
            error_details['truncated_column'] = truncation_match.group(1)
        
        type_error_match = re.search(r'Conversion failed.*?column \'([^\']+)\'', error_str_full, re.IGNORECASE)
        if type_error_match:
            error_details['type_error_column'] = type_error_match.group(1)
        
        # Build comprehensive error message
        print(f"\n{'='*80}")
        print(f"UPLOAD ERROR - Detailed Information")
        print(f"{'='*80}")
        print(f"\nTable: {table}")
        print(f"Total rows attempted: {len(data)}")
        
        # Show error details
        print(f"\n--- Error Information ---")
        if error_details.get('sql_error_code'):
            print(f"SQL Server Error Code: {error_details['sql_error_code']}")
        if error_details.get('error_code'):
            print(f"ODBC Error Code: {error_details['error_code']}")
        if error_details.get('line_number'):
            print(f"SQL Line Number: {error_details['line_number']}")
        if error_details.get('truncated_column'):
            print(f"⚠ TRUNCATION ERROR in column: '{error_details['truncated_column']}'")
        if error_details.get('type_error_column'):
            print(f"⚠ TYPE CONVERSION ERROR in column: '{error_details['type_error_column']}'")
        
        # Show the actual error message(s) - try multiple sources
        print(f"\nError Message:")
        error_to_display = None
        
        # Priority 1: Extracted error message
        if error_details.get('error_message'):
            error_to_display = error_details['error_message']
        # Priority 2: Full error string
        elif error_str_full and error_str_full != str(e) and len(error_str_full) > 5:
            error_to_display = error_str_full
        # Priority 3: Basic string representation
        elif error_msg and len(error_msg) > 5:
            error_to_display = error_msg
        # Priority 4: Exception args
        elif hasattr(e, 'args') and len(e.args) > 0:
            error_to_display = str(e.args)
        
        if error_to_display:
            # Format multi-line errors nicely
            error_lines = str(error_to_display).split('\n')
            for line in error_lines[:10]:  # Limit to first 10 lines
                print(f"  {line}")
            if len(error_lines) > 10:
                print(f"  ... ({len(error_lines) - 10} more lines)")
        else:
            print(f"  (Error message could not be extracted)")
        
        # Always show full exception details as fallback
        print(f"\n--- Full Exception Details (for debugging) ---")
        print(f"  Exception Type: {type(e).__name__}")
        print(f"  Exception String: {str(e)}")
        if hasattr(e, 'args') and e.args:
            print(f"  Exception Args ({len(e.args)}):")
            for i, arg in enumerate(e.args):
                print(f"    [{i}]: {repr(arg)}")
        if hasattr(e, '__cause__') and e.__cause__:
            print(f"  Caused by: {type(e.__cause__).__name__}: {e.__cause__}")
        if hasattr(e, '__context__') and e.__context__:
            print(f"  Context: {type(e.__context__).__name__}: {e.__context__}")
        
        # Show table schema
        print(f"\n--- Table Schema ---")
        if table_cols:
            for col_name, sql_type, max_len in table_cols:
                max_info = f" (max: {max_len})" if max_len else ""
                print(f"  [{col_name}] {sql_type}{max_info}")
        else:
            print("  (Schema information not available)")
        
        # Show DataFrame columns and types
        print(f"\n--- DataFrame Information ---")
        print(f"Columns in DataFrame ({len(cols)}):")
        for i, col in enumerate(cols):
            # Try to get the data type of the first non-null value
            dtype_info = ""
            if not df.empty:
                sample_val = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
                if sample_val is not None:
                    dtype_info = f" (sample type: {type(sample_val).__name__})"
            print(f"  {i+1}. {col}{dtype_info}")
        
        # Show sample data with better formatting
        print(f"\n--- Sample Data (First 3 Rows) ---")
        for i, row in enumerate(data[:3]):
            print(f"\n  Row {i+1}:")
            for j, val in enumerate(row):
                col_name = cols[j] if j < len(cols) else f"Column{j}"
                # Truncate long values for display
                val_str = str(val)
                if len(val_str) > 100:
                    val_str = val_str[:97] + "..."
                val_type = type(val).__name__
                print(f"    [{col_name}]: {val_str} ({val_type})")
        
        # Try to identify potential problematic values
        print(f"\n--- Potential Issues ---")
        issues_found = []
        
        # Check for type mismatches
        if table_cols:
            for col_name, sql_type, max_len in table_cols:
                if col_name in df.columns:
                    # Check for string length issues
                    if max_len and 'char' in sql_type.lower():
                        long_values = df[df[col_name].astype(str).str.len() > max_len]
                        if not long_values.empty:
                            issues_found.append(f"Column '{col_name}' has {len(long_values)} values exceeding max length {max_len}")
                    
                    # Check for NULL in NOT NULL columns (if we had that info)
                    null_count = df[col_name].isna().sum()
                    if null_count > 0:
                        issues_found.append(f"Column '{col_name}' has {null_count} NULL/empty values")
        
        if issues_found:
            for issue in issues_found:
                print(f"  ⚠ {issue}")
        else:
            print("  (Run with more verbose logging to see detailed type information)")
        
        print(f"\n{'='*80}\n")
        raise


def resolve_folder_path(folder: str, base: Path) -> Path:
    """
    Intelligently resolve a folder path:
    1. If absolute and exists, use it directly
    2. If relative, try relative to project root first
    3. Try case-insensitive search within project
    4. Extract just the folder name and search within inbound/
    5. Create the folder if it doesn't exist
    
    Returns the resolved Path (created if necessary).
    """
    p = Path(folder)
    
    # If absolute path exists, use it
    if p.is_absolute() and p.exists():
        return p
    
    # If relative, resolve from project root
    if not p.is_absolute():
        resolved = base / folder
        if resolved.exists():
            return resolved
        
        # Get the folder name (last component: "ActiveInsurance" from "inbound/ActiveInsurance")
        folder_name = Path(folder).name
        
        # Try case-insensitive search within inbound/ specifically first
        inbound_dir = base / 'inbound'
        if inbound_dir.exists():
            try:
                for existing_path in inbound_dir.iterdir():
                    if existing_path.is_dir() and existing_path.name.lower() == folder_name.lower():
                        return existing_path
            except Exception:
                pass
        
        # Try case-insensitive search within the entire project
        for existing_path in base.rglob(folder_name):
            if existing_path.is_dir() and existing_path.name.lower() == folder_name.lower():
                return existing_path
        
        # If none found, create the folder at the intended path
        resolved.mkdir(parents=True, exist_ok=True)
        return resolved
    
    # For absolute paths that don't exist, create them
    p.mkdir(parents=True, exist_ok=True)
    return p


def find_files(folder: str, patterns, base: Path = None):
    """Find files matching patterns in folder. Auto-resolves and creates folder if needed."""
    if base is None:
        base = Path(__file__).parent.resolve()
    p = resolve_folder_path(folder, base)
    files = []
    for pat in patterns:
        files.extend(sorted(p.glob(pat)))
    return files


def ensure_folders_from_config(cfg_path: Path, base: Path):
    """Create inbound folders referenced in config.json. If folder is relative, create under base/inbound."""
    cfg = json.load(open(cfg_path, 'r', encoding='utf-8'))
    for entry in cfg.get('folders', []):
        folder = entry.get('folder')
        if not folder:
            continue
        p = Path(folder)
        if not p.is_absolute():
            p = base / folder
        p.mkdir(parents=True, exist_ok=True)


def parse_table_name(full_name: str):
    # Accept formats: [db].[schema].[table] or schema.table or table
    parts = [p.strip().strip('[]') for p in re.split(r'\.', full_name) if p.strip()]
    if len(parts) == 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return None, parts[0], parts[1]
    if len(parts) == 1:
        return None, 'dbo', parts[0]
    raise ValueError(f'Unable to parse table name: {full_name}')


def get_table_columns(conn, full_table_name: str):
    db, schema, table = parse_table_name(full_table_name)
    if db:
        prefix = f"[{db}]."
    else:
        prefix = ''
    
    # Query actual column sizes from sys.columns which stores the max_length attribute
    sql = f"""
        SELECT 
            c.name as COLUMN_NAME,
            t.name as DATA_TYPE,
            c.max_length as CHAR_MAX
        FROM {prefix}sys.columns c
        INNER JOIN {prefix}sys.types t ON c.user_type_id = t.user_type_id
        INNER JOIN {prefix}sys.tables tbl ON c.object_id = tbl.object_id
        INNER JOIN {prefix}sys.schemas s ON tbl.schema_id = s.schema_id
        WHERE s.name = ? AND tbl.name = ?
        ORDER BY c.column_id
    """
    cur = conn.cursor()
    cur.execute(sql, (schema, table))
    cols = []
    
    for row in cur.fetchall():
        col_name = row.COLUMN_NAME
        data_type = row.DATA_TYPE
        char_max = row.CHAR_MAX
        
        # sys.columns stores max_length in bytes
        # For nvarchar, max_length is 2 * actual char count, for varchar it's 1:1
        if char_max and char_max > 0:
            if 'nvarchar' in data_type.lower():
                char_max = char_max // 2  # Convert bytes to characters
            elif 'varchar' in data_type.lower() or 'char' in data_type.lower():
                char_max = char_max  # Already in characters
        
        cols.append((col_name, data_type, char_max if char_max and char_max > 0 else None))
    
    return cols


def sql_type_to_coercion(data_type: str):
    t = data_type.lower()
    if t in ('int', 'smallint', 'bigint', 'tinyint'):
        return 'int'
    if t in ('decimal', 'numeric', 'money', 'smallmoney', 'float', 'real'):
        return 'float'
    if t in ('bit',):
        return 'bool'
    if 'char' in t or 'text' in t or 'xml' in t or 'uniqueidentifier' in t:
        return 'str'
    if 'date' in t or 'time' in t or 'datetime' in t or 'smalldatetime' in t:
        return 'datetime'
    # default to string
    return 'str'


def prepare_dataframe_for_table(df: 'pd.DataFrame', table_cols, filename=None):
    """Align and coerce a DataFrame to the target table columns.
    table_cols: list of (colname, data_type, char_max_length)
    Returns the coerced DataFrame ordered to match table_cols.
    Uses fuzzy matching to handle minor spelling variations in column names.
    Truncates string values that exceed column width limits.
    """
    # normalize column names for matching (case-insensitive)
    orig_cols = list(df.columns)
    col_map = {c.lower().strip(): c for c in orig_cols}
    prepared = df.copy()
    prepared.columns = [c.strip() for c in prepared.columns]

    # Fuzzy match function
    def fuzzy_match(expected, available, threshold=0.85):
        """Find best fuzzy match for expected column name in available columns."""
        if SequenceMatcher is None:
            return None
        best_match = None
        best_ratio = threshold
        for avail in available:
            ratio = SequenceMatcher(None, expected.lower(), avail.lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = avail
        return best_match

    # Ensure all expected columns exist
    expected_cols = [c for c, _, _ in table_cols]
    missing_cols = []
    extra_cols = []
    
    for exp in expected_cols:
        # try exact case-insensitive match first
        key = exp.lower()
        if key in col_map:
            actual = col_map[key]
            if actual != exp:
                prepared.rename(columns={actual: exp}, inplace=True)
        else:
            # try fuzzy match (e.g., "Prim. Oncologist" -> "Prim# Oncologist")
            fuzzy = fuzzy_match(exp, prepared.columns)
            if fuzzy:
                print(f"Note: Auto-mapping Excel column '{fuzzy}' to expected '{exp}'")
                prepared.rename(columns={fuzzy: exp}, inplace=True)
            else:
                # missing column -> create with NULLs
                missing_cols.append(exp)
                prepared[exp] = None

    # Track unexpected extra columns (those that didn't match anything)
    for c in list(prepared.columns):
        if c.lower() not in [ec.lower() for ec in expected_cols]:
            extra_cols.append(c)

    # Drop unexpected extra columns
    if extra_cols:
        print(f"Warning: dropping extra columns from file {filename or ''}: {extra_cols}")
        prepared.drop(columns=extra_cols, inplace=True)
    
    if missing_cols:
        print(f"Warning: Excel missing columns {missing_cols}, will insert NULLs for file {filename or ''}")

    # Reorder to expected
    prepared = prepared[expected_cols]

    # Coerce types and truncate strings to fit column width
    for col_name, sql_type, char_max_length in table_cols:
        coercion = sql_type_to_coercion(sql_type)
        if coercion == 'int':
            prepared[col_name] = pd.to_numeric(prepared[col_name], errors='coerce').astype('Int64')
        elif coercion == 'float':
            prepared[col_name] = pd.to_numeric(prepared[col_name], errors='coerce').astype('float64')
        elif coercion == 'bool':
            # Handle BIT columns: convert various formats to boolean, empty/blank to NULL
            def convert_to_bit(v):
                # Handle None, NaN, empty strings, whitespace - convert to NULL
                if pd.isna(v):
                    return pd.NA
                if isinstance(v, str) and v.strip() == '':
                    return pd.NA
                # Handle boolean values directly
                if isinstance(v, bool):
                    return v
                # Handle numeric: 1 = True, 0 = False, anything else = NULL
                if isinstance(v, (int, float)):
                    if v == 1:
                        return True
                    elif v == 0:
                        return False
                    else:
                        return pd.NA
                # Handle string representations (case-insensitive)
                v_str = str(v).strip().lower()
                if v_str in ('1', 'true', 'yes', 'y', 't', 'on'):
                    return True
                elif v_str in ('0', 'false', 'no', 'n', 'f', 'off', ''):
                    return False
                # Default to NULL for unrecognized values (empty strings already handled above)
                return pd.NA
            
            prepared[col_name] = prepared[col_name].apply(convert_to_bit).astype('boolean')
        elif coercion == 'datetime':
            prepared[col_name] = pd.to_datetime(prepared[col_name], errors='coerce')
        else:
            # String type - convert to string and truncate if needed
            prepared[col_name] = prepared[col_name].astype('string')
            # Truncate strings to fit column width if a limit is defined
            if char_max_length and char_max_length > 0:
                prepared[col_name] = prepared[col_name].apply(
                    lambda x: str(x)[:char_max_length] if pd.notna(x) else None
                )

    return prepared


def validate_and_prepare_files_for_entry(conn, entry, base: Path):
    """Given a config entry, find files in the folder, convert non-Excel to Excel, align columns and coerce types.
    Returns list of tuples (original_path, prepared_df)
    """
    folder = entry.get('folder')
    if not folder:
        return []
    
    # Use the smart path resolver
    fpath = resolve_folder_path(folder, base)
    patterns = entry.get('file_patterns', ['*.xlsx', '*.xls'])
    files = find_files(folder, patterns, base)

    prepared_list = []
    if not files:
        return prepared_list

    # acquire table schema
    ttable = entry.get('target_table')
    if not ttable:
        return prepared_list
    table_cols = get_table_columns(conn, ttable)

    for f in files:
        try:
            # read file (Excel expected)
            if f.suffix.lower() in ('.xls', '.xlsx'):
                try:
                    df = pd.read_excel(f)
                except Exception as e:
                    # Provide a clearer message for common tempfile/permission issues
                    msg = str(e)
                    if isinstance(e, (PermissionError, OSError)) or 'temp' in msg.lower() or 'temporary' in msg.lower():
                        print(f"Error reading Excel file {f}: {e}\n"
                              "This often indicates a problem creating temporary files (permissions, disk space, or OneDrive sync)."
                              " Try freeing disk space, checking permissions on your TEMP directory, or disabling OneDrive sync for this folder.")
                    else:
                        print(f"Error reading Excel file {f}: {e}")
                    continue
            else:
                # attempt to read as CSV and convert to DataFrame
                try:
                    df = pd.read_csv(f)
                except Exception:
                    print(f"Skipping unreadable file {f}")
                    continue

            # prepare DataFrame
            df_prepared = prepare_dataframe_for_table(df, table_cols, filename=str(f.name))
            prepared_list.append((f, df_prepared, table_cols))
        except Exception as e:
            print(f"Error preparing file {f}: {e}")
    return prepared_list


def upload_from_folders(cfg_path: Path):
    if pd is None:
        raise SystemExit('pandas is not installed. Install with: pip install pandas openpyxl')
    cfg = json.load(open(cfg_path, 'r', encoding='utf-8'))
    base_dir = Path(__file__).parent.resolve()

    # First scan all configured folders to see if any matching files exist.
    any_files_found = False
    for entry in cfg.get('folders', []):
        folder = entry.get('folder')
        if not folder:
            continue
        patterns = entry.get('file_patterns', ['*.xlsx', '*.xls'])
        files = find_files(folder, patterns, base_dir)
        if files:
            any_files_found = True
            break

    if not any_files_found:
        print('No files found in any configured inbound folders. Skipping upload step.')
        return

    # Proceed to connect and upload only if files were found
    conn = connect_from_cfg(cfg['db'])
    try:
        for entry in cfg.get('folders', []):
            ttable = entry.get('target_table')
            folder = entry.get('folder')
            if not ttable:
                print(f"Skipping folder {folder} (no target_table set).")
                continue
            if not folder:
                print('No folder set for entry, skipping')
                continue

            prepared_files = validate_and_prepare_files_for_entry(conn, entry, base_dir)
            if not prepared_files:
                print(f"No valid files found in {folder}")
                continue
            
            # Determine upload mode from config
            upload_mode = entry.get('upload_mode', 'append')  # default to append for backward compatibility
            
            for orig_path, df, table_cols in prepared_files:
                print(f"Uploading {orig_path} -> {ttable} (mode: {upload_mode})")
                upload_df_to_table(conn, df, ttable, upload_mode=upload_mode, table_cols=table_cols)
                print(f"Uploaded {len(df)} rows from {orig_path.name}")
    finally:
        conn.close()


def run_sql_scripts(seq_files, cfg_path: Path):
    cfg = json.load(open(cfg_path, 'r', encoding='utf-8'))
    conn = connect_from_cfg(cfg['db'])
    try:
        cursor = conn.cursor()
        for sqlfile in seq_files:
            print('Running', sqlfile)
            text = open(sqlfile, 'r', encoding='utf-8-sig').read()
            for batch in split_sql_batches(text):
                cursor.execute(batch)
        conn.commit()
    finally:
        conn.close()


def main():
    base = Path(__file__).parent.resolve()
    inbound = base / 'inbound'
    cfg_path = base / 'config.json'

    p = argparse.ArgumentParser(description='Upload files and run SQL scripts in order')
    p.add_argument('--init', action='store_true', help='Create inbound folders and template config.json')
    p.add_argument('--config', default=str(cfg_path), help='Path to config.json')
    p.add_argument('--upload', action='store_true', help='Upload files from configured folders')
    p.add_argument('--run-sql', action='store_true', help='Execute SQL scripts (in alphabetical/numeric order)')
    p.add_argument('--test-connection', action='store_true', help='Test DB connection and report server/user info')
    p.add_argument('--list-tables', action='store_true', help='List accessible base tables via INFORMATION_SCHEMA')
    args = p.parse_args()

    sql_files = list_sql_files(base)
    if not sql_files:
        print('No .sql files found in current directory.')
        return

    if args.test_connection:
        if not Path(args.config).exists():
            print('config.json not found. Run --init first.')
            return
        rc = test_connection(Path(args.config))
        sys.exit(rc)

    if args.list_tables:
        if not Path(args.config).exists():
            print('config.json not found. Run --init first.')
            return
        rc = list_tables(Path(args.config))
        sys.exit(rc)

    if args.init:
        # Create inbound folders from existing config if present, otherwise create default numbered folders
        if Path(args.config).exists():
            ensure_folders_from_config(Path(args.config), base)
            print('Created inbound folders from config.json')
        else:
            folders = create_inbound_dirs(sql_files, inbound)
            write_template_config(sql_files, folders, Path(args.config))
        return

    if args.upload:
        if not Path(args.config).exists():
            print('config.json not found. Run --init first.')
            return
        upload_from_folders(Path(args.config))

    if args.run_sql:
        # Always attempt to upload files first using the configured folders
        if Path(args.config).exists():
            try:
                print('Uploading files from configured inbound folders before running SQL scripts...')
                ensure_folders_from_config(Path(args.config), base)
                upload_from_folders(Path(args.config))
            except Exception as e:
                print('Warning: upload step failed or skipped:', e)
        run_sql_scripts([str((base / f)) for f in sql_files], Path(args.config))


if __name__ == '__main__':
    main()
